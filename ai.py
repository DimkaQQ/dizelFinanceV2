# -*- coding: utf-8 -*-
"""
DizelFinance — AI модуль (Claude API через прокси)
"""

import json
import base64
import logging
import time
import re
import requests

from config import (
    ANTHROPIC_API_KEY, CLAUDE_MODEL, CLAUDE_PROXY,
    INCOME_CATS, EXPENSE_CATS, ASSET_CATS, ALL_CATEGORIES,
    is_julian, is_internal_transfer,
)

log = logging.getLogger(__name__)

from datetime import datetime, timedelta

# ══════════════════════════════════════════════════════════════════════════════
# Нормализация дат
# ══════════════════════════════════════════════════════════════════════════════

def normalize_date(date_str: str) -> str:
    """
    Нормализует дату. Если нет/неверная — возвращает сегодняшнюю.
    """
    today = datetime.now()
    
    if not date_str:
        return today.strftime("%d.%m.%Y")
    
    # Убираем время если есть
    date_part = date_str.strip().split(",")[0].strip()
    
    # Проверяем на относительные даты
    lower = date_part.lower()
    if lower in ["сегодня", "today", "сейчас", "now", "", "none", "null"]:
        return today.strftime("%d.%m.%Y")
    if lower in ["вчера", "yesterday"]:
        return (today - timedelta(days=1)).strftime("%d.%m.%Y")
    
    # Пробуем разные форматы
    for fmt in ["%d.%m.%Y", "%d.%m", "%Y-%m-%d", "%d %B %Y", "%d.%m.%y"]:
        try:
            parsed = datetime.strptime(date_part, fmt)
            # Если только день.месяц — добавляем текущий год
            if fmt == "%d.%m" or "%d.%m.%y" in fmt:
                parsed = parsed.replace(year=today.year)
            # Валидация: не раньше 2020, не позже +30 дней (для будущих оплат)
            if parsed.year < 2020 or parsed > today + timedelta(days=30):
                return today.strftime("%d.%m.%Y")
            return parsed.strftime("%d.%m.%Y")
        except ValueError:
            continue
    
    # Ни один формат не подошёл — возвращаем сегодня
    return today.strftime("%d.%m.%Y")

# ══════════════════════════════════════════════════════════════════════════════
# Базовый запрос к Claude
# ══════════════════════════════════════════════════════════════════════════════

def ask_claude(prompt: str, image_bytes: bytes = None,
               mime_type: str = "image/jpeg") -> str:
    content = []
    if image_bytes:
        content.append({
            "type": "image",
            "source": {
                "type": "base64",
                "media_type": mime_type,
                "data": base64.b64encode(image_bytes).decode(),
            }
        })
    content.append({"type": "text", "text": prompt})

    payload = {
        "model": CLAUDE_MODEL,
        "max_tokens": 8000,
        "messages": [{"role": "user", "content": content}],
    }
    headers = {
        "x-api-key": ANTHROPIC_API_KEY,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json",
    }

    for attempt in range(5):
        try:
            r = requests.post(CLAUDE_PROXY, json=payload, headers=headers, timeout=90)
            if r.status_code == 200:
                return r.json()["content"][0]["text"].strip()
            log.warning(f"Claude {r.status_code}: {r.text[:200]}")
        except Exception as e:
            log.error(f"Claude attempt {attempt+1}: {e}")
        time.sleep(4 * (attempt + 1))
    raise RuntimeError("Claude недоступен после 5 попыток")

# ══════════════════════════════════════════════════════════════════════════════
# Парсинг JSON из ответа
# ══════════════════════════════════════════════════════════════════════════════

def extract_json(text: str):
    text = re.sub(r'```json\s*', '', text)
    text = re.sub(r'```\s*', '', text).strip()

    def find_end(s, start, op, cl):
        depth, in_str, esc = 0, False, False
        for i in range(start, len(s)):
            c = s[i]
            if esc:       esc = False; continue
            if c == '\\' and in_str: esc = True; continue
            if c == '"':  in_str = not in_str; continue
            if in_str:    continue
            if c == op:   depth += 1
            elif c == cl: depth -= 1
            if depth == 0: return i
        return -1

    for ch, end_ch in [('[', ']'), ('{', '}')]:
        pos = text.find(ch)
        if pos != -1:
            end = find_end(text, pos, ch, end_ch)
            if end != -1:
                try:
                    return json.loads(text[pos:end+1])
                except json.JSONDecodeError:
                    pass
    return json.loads(text)

# ══════════════════════════════════════════════════════════════════════════════
# Определение категории для одной транзакции
# ══════════════════════════════════════════════════════════════════════════════

def resolve_category(category: str, tx_type: str) -> tuple[str, str]:
    """Возвращает (category, section) — гарантированно из ALL_CATEGORIES."""
    if category in ALL_CATEGORIES:
        return category, ALL_CATEGORIES[category]
    # фоллбэк
    if tx_type == "Доход":
        return "Прочее (доход)", "Доходы"
    if tx_type == "Актив":
        return "Инвестиции в консалтинг", "Движение активов"
    return "Прочее (рег)", "Регулярные расходы"

def guess_category(merchant: str, amount: float,
                   tx_type: str = "Расход", hint: str = "",
                   history_text: str = "") -> tuple[str, str]:
    """Возвращает (category, section)."""
    if tx_type == "Расход" and is_julian(merchant):
        return "Джулиан", "Регулярные расходы"
    if is_internal_transfer(merchant, hint):
        return "Инвестиции в консалтинг", "Движение активов"

    cats = INCOME_CATS if tx_type == "Доход" else (
        ASSET_CATS if tx_type == "Актив" else EXPENSE_CATS
    )
    hist = f"История:\n{history_text}\n" if history_text else ""
    hint_line = f"Подсказка банка: «{hint}»\n" if hint else ""

    prompt = (
        f"Определи категорию транзакции.\n"
        f"Место: {merchant} | Сумма: {amount} | Тип: {tx_type}\n"
        f"{hint_line}{hist}"
        f"Категории: {json.dumps(cats, ensure_ascii=False)}\n"
        f'Ответь ТОЛЬКО JSON: {{"category": "название"}}\n'
        f"Правила:\n"
        f"- Только из предложенного списка\n"
        f"- Ветеринар/зоомагазин → 'Джулиан'\n"
        f"- Рестораны/кафе → 'Рестораны/кафе/фастфуд'\n"
        f"- Если место есть в истории — используй ту же категорию"
    )
    try:
        result = ask_claude(prompt)
        data = extract_json(result)
        if isinstance(data, dict):
            return resolve_category(data.get("category", ""), tx_type)
    except Exception as e:
        log.error(f"guess_category: {e}")
    return resolve_category("", tx_type)

def guess_categories_batch(transactions: list[dict],
                            history_text: str = "") -> list[tuple[str, str]]:
    """Батч-угадывание для списка транзакций."""
    if not transactions:
        return []

    pre: dict[int, tuple] = {}
    ai_idx: list[int] = []

    for i, tx in enumerate(transactions):
        m = tx.get("merchant", "")
        h = tx.get("category_hint", "")
        t = tx.get("tx_type", "Расход")
        if t == "Расход" and is_julian(m):
            pre[i] = ("Джулиан", "Регулярные расходы")
        elif is_internal_transfer(m, h):
            pre[i] = ("Инвестиции в консалтинг", "Движение активов")
        else:
            ai_idx.append(i)

    if not ai_idx:
        return [pre[i] for i in range(len(transactions))]

    items = []
    for i in ai_idx:
        tx = transactions[i]
        h = tx.get("category_hint", "")
        items.append(
            f'{i}: merchant="{tx.get("merchant","")}", '
            f'amount={tx.get("amount",0)}, '
            f'type="{tx.get("tx_type","Расход")}"'
            + (f', hint="{h}"' if h else "")
        )

    hist = f"История:\n{history_text}\n" if history_text else ""
    prompt = (
        f"Определи категорию для каждой транзакции.\n"
        f"Транзакции:\n{chr(10).join(items)}\n"
        f"Категории расходов: {json.dumps(EXPENSE_CATS, ensure_ascii=False)}\n"
        f"Категории доходов: {json.dumps(INCOME_CATS, ensure_ascii=False)}\n"
        f"Категории активов: {json.dumps(ASSET_CATS, ensure_ascii=False)}\n"
        f"{hist}"
        f'Ответь ТОЛЬКО JSON массивом: [{{"index": 0, "category": "название"}}, ...]\n'
        f"Правила:\n"
        f"- Только из предложенных категорий\n"
        f"- type=Расход → из расходов, type=Доход → из доходов, type=Актив → из активов\n"
        f"- Ветеринар/зоо → 'Джулиан', рестораны/кафе → 'Рестораны/кафе/фастфуд'"
    )

    ai_results: dict[int, tuple] = {}
    try:
        raw = extract_json(ask_claude(prompt))
        if isinstance(raw, list):
            for item in raw:
                idx = item.get("index", -1)
                if idx in ai_idx:
                    t = transactions[idx].get("tx_type", "Расход")
                    ai_results[idx] = resolve_category(item.get("category", ""), t)
    except Exception as e:
        log.error(f"guess_categories_batch: {e}")

    # фоллбэк для пропущенных
    for i in ai_idx:
        if i not in ai_results:
            t = transactions[i].get("tx_type", "Расход")
            ai_results[i] = resolve_category("", t)

    return [pre.get(i) or ai_results.get(i, ("Прочее (рег)", "Регулярные расходы"))
            for i in range(len(transactions))]

# ══════════════════════════════════════════════════════════════════════════════
# Парсинг файлов
# ══════════════════════════════════════════════════════════════════════════════

_EXTRACT_PROMPT = """Это {source}. Извлеки ВСЕ транзакции.
Для каждой верни:
- date: ДД.ММ.ГГГГ (если "Сегодня"/"Today"/даты нет — верни пустую строку "")
- amount: число (положительное, без пробелов и валюты, например 21600.00)
- currency: RUB/USD/EUR/KZT/IDR/VND (точно из этого списка)
- merchant: название места (коротко, без адреса и деталей)
- tx_type: "Расход" или "Доход" (Расход = списание, Доход = зачисление)
- category_hint: категория если видна на скрине, иначе ""
Игнорируй: переводы между своими счетами, балансы, "доступно", "лимит", комиссии.
Ответ ТОЛЬКО чистым JSON массивом. Если нет транзакций — [].
Пример: [{{"date":"01.01.2025","amount":1500,"currency":"RUB","merchant":"Пятёрочка","tx_type":"Расход","category_hint":"еда"}}]"""

def parse_screenshot(image_bytes: bytes, mime_type: str = "image/jpeg") -> list[dict]:
    prompt = _EXTRACT_PROMPT.format(source="скриншот банковского приложения")
    try:
        result = ask_claude(prompt, image_bytes=image_bytes, mime_type=mime_type)
        data = extract_json(result)
        if isinstance(data, list): 
            transactions = data
        elif isinstance(data, dict): 
            transactions = [data]
        else: 
            transactions = []
        
                # ❗ Нормализуем даты и валидируем поля
        for tx in transactions:
            # Нормализация даты
            raw_date = tx.get("date", "")
            tx["date"] = normalize_date(raw_date) + ", 12:00"
            
            # Гарантируем что amount — число
            try:
                amt = tx.get("amount")
                if isinstance(amt, str):
                    amt = amt.replace(" ", "").replace(",", ".").replace("Rp", "").strip()
                tx["amount"] = float(amt) if amt else 0.0
            except:
                tx["amount"] = 0.0
            
            # Гарантируем валюту
            if tx.get("currency") not in CURRENCIES:
                tx["currency"] = "RUB"
            
            # Гарантируем tx_type
            if tx.get("tx_type") not in ["Расход", "Доход", "Актив"]:
                # Определяем по знаку суммы если есть
                if tx.get("amount", 0) < 0:
                    tx["tx_type"] = "Расход"
                else:
                    tx["tx_type"] = "Расход"  # дефолт
            
            # Гарантируем наличие merchant
            if not tx.get("merchant"):
                tx["merchant"] = "Неизвестно"
        
        return transactions
    except Exception as e:
        log.error(f"parse_screenshot: {e}")
    return []

def parse_pdf(pdf_bytes: bytes) -> list[dict]:
    try:
        import fitz
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        all_tx = []
        for i, page in enumerate(doc):
            pix = page.get_pixmap(matrix=fitz.Matrix(150/72, 150/72))
            img = pix.tobytes("png")
            prompt = _EXTRACT_PROMPT.format(source=f"страница {i+1} банковской выписки")
            try:
                result = ask_claude(prompt, image_bytes=img, mime_type="image/png")
                data = extract_json(result)
                if isinstance(data, list): 
                    all_tx.extend(data)
                elif isinstance(data, dict): 
                    all_tx.append(data)
            except Exception as e:
                log.error(f"parse_pdf page {i}: {e}")
        doc.close()
        
               # ❗ Нормализуем даты и валидируем поля
        for tx in transactions:
            # Нормализация даты
            raw_date = tx.get("date", "")
            tx["date"] = normalize_date(raw_date) + ", 12:00"
            
            # Гарантируем что amount — число
            try:
                amt = tx.get("amount")
                if isinstance(amt, str):
                    amt = amt.replace(" ", "").replace(",", ".").replace("Rp", "").strip()
                tx["amount"] = float(amt) if amt else 0.0
            except:
                tx["amount"] = 0.0
            
            # Гарантируем валюту
            if tx.get("currency") not in CURRENCIES:
                tx["currency"] = "RUB"
            
            # Гарантируем tx_type
            if tx.get("tx_type") not in ["Расход", "Доход", "Актив"]:
                # Определяем по знаку суммы если есть
                if tx.get("amount", 0) < 0:
                    tx["tx_type"] = "Расход"
                else:
                    tx["tx_type"] = "Расход"  # дефолт
            
            # Гарантируем наличие merchant
            if not tx.get("merchant"):
                tx["merchant"] = "Неизвестно"
        
        return all_tx
    except ImportError:
        log.error("fitz не установлен")
    except Exception as e:
        log.error(f"parse_pdf: {e}")
    return []

def parse_xlsx(file_bytes: bytes) -> list[dict]:
    try:
        import openpyxl
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        ws = wb.active
        header_row = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if any("дата операции" in str(c).lower() for c in row if c):
                header_row = i
                break
        if header_row is None:
            log.error("XLSX: заголовок не найден")
            return []
        rows = list(ws.iter_rows(values_only=True))
        txs = []
        for row in rows[header_row + 1:]:
            cells = [str(c).replace('\xa0',' ').strip() if c else '' for c in row]
            if len(cells) < 13: continue
            if not re.match(r'\d{2}\.\d{2}\.\d{4}', cells[0]): continue
            amt_str = cells[12].replace(' ', '').replace(',', '.')
            if not amt_str: continue
            try:
                amt = float(amt_str)
            except ValueError:
                continue
            m = re.search(r'место совершения операции:[^,]+/([^,\n]+?)(?:,|\s+MCC|$)',
                          cells[11], re.I)
            merchant = m.group(1).strip() if m else cells[11][:50]
            txs.append({
                "date": cells[0],
                "amount": abs(amt),
                "currency": "RUB",
                "merchant": merchant,
                "tx_type": "Доход" if amt > 0 else "Расход",
                "category_hint": cells[4],
            })
        return txs
    except Exception as e:
        log.error(f"parse_xlsx: {e}")
    return []

def parse_sms(text: str) -> dict | None:
    prompt = (
        f"Извлеки данные транзакции из SMS:\n{text}\n"
        f'Ответь ТОЛЬКО JSON: {{"amount":0.0,"currency":"RUB","merchant":"...","tx_type":"Расход","date":"ДД.ММ.ГГГГ или пусто"}}\n'
        f"Если это НЕ транзакция — верни {{\"error\":\"not_transaction\"}}"
    )
    try:
        result = ask_claude(prompt)
        data = extract_json(result)
        if isinstance(data, dict) and data.get("error") == "not_transaction":
            return None
        return data if isinstance(data, dict) else None
    except Exception as e:
        log.error(f"parse_sms: {e}")
    return None

# Добавь этот код в ai.py (после parse_sms)

_SUMMARY_PROMPT = """
Это скриншот аналитики или статистики (сводка по категориям).
Здесь НЕТ отдельных операций. Извлеки ИТОГОВУЮ сумму для КАЖДОЙ категории.

Для каждой категории верни:
- merchant: Название категории (например "Рестораны")
- amount: Общая сумма (число, без пробелов и валюты)
- currency: RUB или USD/EUR и т.д. (если не указано, предполагай RUB)
- tx_type: "Расход" (если это траты) или "Доход" (если доходы)

Верни ТОЛЬКО чистый JSON массив. Если данных нет — [].
Пример: [{"merchant": "Рестораны", "amount": 30000, "currency": "RUB", "tx_type": "Расход"}]
"""

def parse_category_summary(image_bytes: bytes, mime_type: str = "image/jpeg") -> list[dict]:
    """Парсит скриншоты аналитики (категории с итогами)."""
    try:
        result = ask_claude(_SUMMARY_PROMPT, image_bytes=image_bytes, mime_type=mime_type)
        data = extract_json(result)
        
        if isinstance(data, list):
            transactions = data
        elif isinstance(data, dict):
            transactions = [data]
        else:
            transactions = []
            
        # Нормализация данных
        today = datetime.now().strftime("%d.%m.%Y, 12:00")
        
        for tx in transactions:
            # Дата = Сегодня (так как это сводка)
            tx["date"] = today
            
            # Сумма
            try:
                amt = tx.get("amount", 0)
                if isinstance(amt, str):
                    amt = amt.replace(" ", "").replace(",", ".")
                tx["amount"] = float(amt)
            except:
                tx["amount"] = 0.0
                
            # Валюта
            if tx.get("currency") not in CURRENCIES:
                tx["currency"] = "RUB"
                
            # Тип (по умолчанию Расход)
            if tx.get("tx_type") not in ["Расход", "Доход", "Актив"]:
                tx["tx_type"] = "Расход"
                
            # Название (если пустое)
            if not tx.get("merchant"):
                tx["merchant"] = "Без категории"
                
            # Очищаем поле hint, так как это уже категория
            tx["category_hint"] = "" 
            
        return transactions
    except Exception as e:
        log.error(f"parse_category_summary: {e}")
        return []